from __future__ import annotations
import os
import json
import tempfile
import logging
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import re
import pandas as pd
import numpy as np

# Google Document AI imports
try:
    from google.cloud import documentai
    from google.oauth2 import service_account
    DOCUMENT_AI_AVAILABLE = True
except ImportError:
    DOCUMENT_AI_AVAILABLE = False
    logging.warning("Google Document AI not available. Install with: pip install google-cloud-documentai")

# Tesseract fallback imports
try:
    import pytesseract
    from PIL import Image
    import cv2
    import numpy as np
    TESSERACT_AVAILABLE = True
except ImportError:
    Image = None  # type: ignore[assignment]
    cv2 = None    # type: ignore[assignment]
    TESSERACT_AVAILABLE = False
    logging.warning("Tesseract OCR not available. Install with: pip install pytesseract opencv-python")

# PDF processing imports
try:
    import fitz  # PyMuPDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logging.warning("PDF processing not available. Install with: pip install PyMuPDF")

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Excel processing imports
try:
    import openpyxl
    import xlrd
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.error("Excel processing libraries not available: No module named 'xlrd'")
    logger.error("Install required libraries: pip install openpyxl xlrd pandas")

class DocumentAIProcessor:
    """Google Document AI processor for OCR extraction"""
    
    def __init__(self):
        self.client = None
        self.processor_id = None
        self.project_id = None
        self.location = None
        self._initialize_client()
    
    def _initialize_client(self):
        """Initialize Document AI client with credentials"""
        try:
            # Try to get credentials from environment or secrets
            credentials_path = os.getenv('GOOGLE_APPLICATION_CREDENTIALS')
            if not credentials_path:
                # Try to get from streamlit secrets
                try:
                    import streamlit as st
                    logger.info("Streamlit imported successfully")

                    # Check if secrets are available
                    if hasattr(st, 'secrets') and st.secrets is not None:
                        logger.info(f"st.secrets keys: {list(st.secrets.keys()) if hasattr(st.secrets, 'keys') else 'No keys method'}")

                        if 'google_documentai' in st.secrets:
                            logger.info("Found google_documentai in st.secrets")
                            docai_config = st.secrets['google_documentai']

                            # Check if credentials are provided directly in secrets
                            if 'type' in docai_config and 'private_key' in docai_config:
                                # Use credentials from st.secrets
                                logger.info("Using credentials from st.secrets")
                                credentials_info = {
                                    'type': docai_config['type'],
                                    'project_id': docai_config['project_id'],
                                    'private_key': docai_config['private_key'],
                                    'client_email': docai_config['client_email'],
                                    'token_uri': docai_config.get('token_uri', 'https://oauth2.googleapis.com/token'),
                                    'auth_uri': docai_config.get('auth_uri', 'https://accounts.google.com/o/oauth2/auth'),
                                }

                            # Add optional fields
                            if 'private_key_id' in docai_config:
                                credentials_info['private_key_id'] = docai_config['private_key_id']
                            if 'client_id' in docai_config:
                                credentials_info['client_id'] = docai_config['client_id']

                            credentials = service_account.Credentials.from_service_account_info(credentials_info)
                            self.client = documentai.DocumentProcessorServiceClient(credentials=credentials)
                            self.project_id = docai_config['project_id']
                            self.processor_id = docai_config['processor_id']
                            self.location = docai_config.get('location', 'us')
                            logger.info("Document AI initialized from st.secrets credentials")
                            return
                except Exception as e:
                    logger.warning(f"Could not load from Streamlit secrets: {e}")
                    # Try to load from the JSON file path specified in st.secrets
                    try:
                        # Check if there's a GOOGLE_APPLICATION_CREDENTIALS path in st.secrets
                        if hasattr(st, 'secrets') and st.secrets is not None and 'google_documentai' in st.secrets:
                            creds_path = st.secrets['google_documentai'].get('GOOGLE_APPLICATION_CREDENTIALS')
                            if creds_path and os.path.exists(creds_path):
                                logger.info(f"Trying to load from st.secrets credentials path: {creds_path}")
                                os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = creds_path
                                creds = service_account.Credentials.from_service_account_file(creds_path)
                                self.client = documentai.DocumentProcessorServiceClient(credentials=creds)

                                # Load project and processor info from the file
                                with open(creds_path, 'r') as f:
                                    creds_data = json.load(f)
                                    self.project_id = creds_data.get('project_id')

                                # Get processor info from st.secrets
                                docai_config = st.secrets['google_documentai']
                                self.processor_id = docai_config.get('processor_id')
                                self.location = docai_config.get('location', 'us')
                                logger.info("Document AI initialized from st.secrets credentials file path")
                                return
                    except Exception as file_error:
                        logger.warning(f"Could not load from st.secrets credentials file: {file_error}")

                    logger.error(f"st.secrets keys: {list(st.secrets.keys()) if hasattr(st.secrets, 'keys') else 'No keys'}")
            
            if credentials_path and os.path.exists(credentials_path):
                # Set credentials path for Google auth
                os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = credentials_path
                creds = service_account.Credentials.from_service_account_file(credentials_path)
                self.client = documentai.DocumentProcessorServiceClient(credentials=creds)

                # Load project details from credentials file
                with open(credentials_path, 'r') as f:
                    creds_data = json.load(f)
                    self.project_id = creds_data.get('project_id')

                # Try to get processor configuration from st.secrets or environment
                try:
                    import streamlit as st
                    if hasattr(st, 'secrets') and 'google_documentai' in st.secrets:
                        docai_config = st.secrets['google_documentai']
                        self.processor_id = docai_config.get('processor_id')
                        self.location = docai_config.get('location', 'us')
                        logger.info("Document AI initialized from credentials file + st.secrets config")
                        return
                except Exception:
                    pass

                # Fallback to environment variables
                self.processor_id = os.getenv('GOOGLE_CLOUD_PROCESSOR_ID', 'ada42b6854568248')
                self.location = os.getenv('GOOGLE_CLOUD_LOCATION', 'us')
                logger.info("Document AI initialized from credentials file + environment config")
            else:
                # Try default environment variables as final fallback
                try:
                    self.client = documentai.DocumentProcessorServiceClient()
                    self.project_id = os.getenv('GOOGLE_CLOUD_PROJECT_ID', 'agriai-cbd8b')
                    self.processor_id = os.getenv('GOOGLE_CLOUD_PROCESSOR_ID', 'ada42b6854568248')
                    self.location = os.getenv('GOOGLE_CLOUD_LOCATION', 'us')
                    logger.info("Document AI initialized with default environment variables")
                except Exception as default_error:
                    logger.warning(f"Could not initialize with default credentials: {default_error}")
                    logger.warning("No Document AI credentials found")

        except Exception as e:
            logger.error(f"Failed to initialize Document AI client: {e}")
            self.client = None
    
    def process_document(self, file_path: str) -> Optional[Dict]:
        """Process document with Google Document AI"""
        if not self.client or not self.processor_id or not self.project_id:
            logger.error("Document AI not properly configured")
            return None
        
        try:
            # Read file content
            with open(file_path, 'rb') as file:
                file_content = file.read()
            
            # Determine MIME type
            file_ext = os.path.splitext(file_path)[1].lower()
            mime_type_map = {
                '.pdf': 'application/pdf',
                '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                '.xls': 'application/vnd.ms-excel',
                '.png': 'image/png',
                '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg'
            }
            mime_type = mime_type_map.get(file_ext, 'application/octet-stream')

            # Create the document
            raw_document = documentai.RawDocument(
                content=file_content,
                mime_type=mime_type
            )
            
            # Configure the process request
            name = f"projects/{self.project_id}/locations/{self.location}/processors/{self.processor_id}"
            request = documentai.ProcessRequest(
                name=name,
                raw_document=raw_document
            )
            
            # Process the document
            result = self.client.process_document(request=request)
            document = result.document
            
            # Extract text and tables
            extracted_data = {
                'text': document.text,
                'tables': [],
                'success': True,
                'method': 'document_ai'
            }
            
            # Process tables from all pages with enhanced detection
            if hasattr(document, 'pages') and document.pages:
                total_tables_found = 0

                for page_num, page in enumerate(document.pages):
                    logger.info(f"Processing page {page_num + 1} of {len(document.pages)}")

                    if hasattr(page, 'tables') and page.tables:
                        logger.info(f"Found {len(page.tables)} tables on page {page_num + 1}")

                        for table_num, table in enumerate(page.tables):
                            logger.info(f"Extracting table {table_num + 1} from page {page_num + 1}")
                            table_data = self._extract_table_data(table, document.text)
                            if table_data:
                                extracted_data['tables'].append(table_data)
                                total_tables_found += 1
                                logger.info(f"Successfully extracted table {table_num + 1} with type: {table_data.get('type', 'unknown')}, samples: {table_data.get('total_samples', 0)}")
                        logger.debug(f"No tables found on page {page_num + 1}")

                        # Try to extract text blocks that might be tabular data
                        if hasattr(page, 'blocks'):
                            logger.debug(f"Page has {len(page.blocks)} blocks, checking for tabular text")
                            text_blocks = self._extract_text_blocks_from_page(page, document.text)
                            if text_blocks:
                                logger.info(f"Found {len(text_blocks)} text blocks on page {page_num + 1}, attempting table extraction")
                                for block_num, block_data in enumerate(text_blocks):
                                    if block_data and block_data.get('type') in ['soil', 'leaf']:
                                        extracted_data['tables'].append(block_data)
                                        total_tables_found += 1
                                        logger.info(f"Successfully extracted text-based table {block_num + 1} with type: {block_data.get('type')}")

                logger.info(f"Total tables extracted: {total_tables_found}")

                # If no tables found but we have document text, try to parse it as tabular data
                if total_tables_found == 0 and document.text:
                    logger.info("No tables found, attempting to parse document text for tabular data")
                    text_tables = self._parse_text_for_tables(document.text)
                    if text_tables:
                        extracted_data['tables'].extend(text_tables)
                        logger.info(f"Successfully extracted {len(text_tables)} tables from text parsing")

            else:
                logger.warning("Document has no pages attribute or no pages")
            
            return extracted_data
            
        except Exception as e:
            logger.error(f"Document AI processing failed: {e}")
            return None
    
    def _extract_table_data(self, table, document_text: str) -> Optional[Dict]:
        """Extract structured data from Document AI table"""
        try:
            rows = []
            headers = []

            # Extract table structure
            for row in table.body_rows:
                row_data = []
                for cell in row.cells:
                    cell_text = self._get_text_from_layout(cell.layout, document_text).strip()
                    row_data.append(cell_text)
                rows.append(row_data)

            # Extract headers if available
            if table.header_rows:
                for cell in table.header_rows[0].cells:
                    header_text = self._get_text_from_layout(cell.layout, document_text).strip()
                    headers.append(header_text)

            if not rows:
                logger.debug("No rows found in table")
                return None

            logger.debug(f"Extracted table with {len(headers)} headers and {len(rows)} rows")

            # Determine table type and structure data accordingly
            table_type = self._determine_table_type(headers, rows)
            
            if table_type == 'soil':
                return self._structure_soil_data(headers, rows)
            elif table_type == 'leaf':
                return self._structure_leaf_data(headers, rows)
            else:
                logger.debug(f"Unknown table type detected with headers: {headers[:3]}...")
                return {
                    'type': 'unknown',
                    'headers': headers,
                    'rows': rows,
                    'samples': []
                }
                    
        except Exception as e:
            logger.error(f"Table extraction failed: {e}")
            return None

    def _extract_text_blocks_from_page(self, page, document_text: str) -> List[Optional[Dict]]:
        """Extract text blocks that might contain tabular data"""
        try:
            text_blocks = []

            if not hasattr(page, 'blocks'):
                return text_blocks

            for block in page.blocks:
                if hasattr(block, 'layout') and hasattr(block, 'text_anchor'):
                    block_text = self._get_text_from_layout(block.layout, document_text)

                    # Check if this block might contain tabular data
                    if self._is_text_block_tabular(block_text):
                        logger.debug(f"Found potentially tabular text block with {len(block_text)} characters")

                        # Try to parse this block as tabular data
                        table_data = self._parse_tabular_text_block(block_text)
                        if table_data:
                            text_blocks.append(table_data)

            return text_blocks

        except Exception as e:
            logger.error(f"Text block extraction failed: {e}")
            return []

    def _is_text_block_tabular(self, text: str) -> bool:
        """Check if a text block contains tabular data"""
        if not text or len(text.strip()) < 50:
            return False

        lines = [line.strip() for line in text.split('\n') if line.strip()]

        # Look for patterns that suggest tabular data
        soil_indicators = ['ph', 'nitrogen', 'organic', 'phosphorus', 'potassium', 'calcium', 'magnesium', 'cec']
        tabular_indicators = ['sample', 's001', 's002', 'mg/kg', 'meq%', '%']

        soil_score = sum(1 for indicator in soil_indicators if indicator.lower() in text.lower())
        tabular_score = sum(1 for indicator in tabular_indicators if indicator.lower() in text.lower())

        # Check for numeric patterns that suggest data columns
        import re
        numeric_patterns = len(re.findall(r'\d+\.?\d*', text))

        return (soil_score >= 2 or tabular_score >= 3) and numeric_patterns >= 5

    def _parse_tabular_text_block(self, text: str) -> Optional[Dict]:
        """Parse a text block that contains tabular data"""
        try:
            lines = [line.strip() for line in text.split('\n') if line.strip()]

            if len(lines) < 2:  # Need at least 2 lines for meaningful parsing
                return None

            logger.debug(f"Parsing text block with {len(lines)} lines")

            # Try different parsing strategies in order of preference
            parsed_tables = []

            # Strategy 1: Look for CSV-like patterns
            if ',' in text or '\t' in text:
                delimiter = ',' if ',' in text else '\t'
                logger.debug(f"Attempting delimiter parsing with '{delimiter}'")
                parsed_tables.extend(self._parse_delimited_text(text, delimiter))

            # Strategy 2: Look for space-aligned columns (more flexible)
            if not parsed_tables and len(lines) >= 3:
                logger.debug("Attempting space-aligned parsing")
                parsed_tables.extend(self._parse_space_aligned_text(lines))

            # Strategy 3: Look for structured patterns with sample IDs (most flexible)
            if not parsed_tables:
                logger.debug("Attempting structured sample text parsing")
                parsed_tables.extend(self._parse_structured_sample_text(lines))

            # Strategy 4: Try to parse as a simple table with any numeric data
            if not parsed_tables and len(lines) >= 2:
                logger.debug("Attempting generic table parsing")
                parsed_tables.extend(self._parse_generic_table(lines))

            # Return the best parsed table
            valid_tables = [table for table in parsed_tables if table and table.get('samples')]
            if valid_tables:
                # Return the table with the most samples
                best_table = max(valid_tables, key=lambda t: len(t.get('samples', [])))
                logger.debug(f"Selected best table with {len(best_table.get('samples', []))} samples of type {best_table.get('type', 'unknown')}")
                return best_table

            return None
            
        except Exception as e:
            logger.error(f"Text block parsing failed: {e}")
            return None

    def _parse_generic_table(self, lines: List[str]) -> List[Optional[Dict]]:
        """Parse a generic table with any numeric data"""
        try:
            tables = []

            # Look for lines that might contain data
            data_lines = []
            for line in lines:
                # Check if line contains numbers (likely data)
                if re.search(r'\d', line):
                    # Split by multiple spaces or tabs
                    parts = re.split(r'\s{2,}|\t', line.strip())
                    if len(parts) >= 2:  # At least 2 columns
                        data_lines.append(parts)

            if len(data_lines) >= 2:  # At least 2 data rows
                # Assume first row is headers, rest are data
                headers = data_lines[0]
                rows = data_lines[1:]

                # Clean headers
                headers = [h.strip() for h in headers if h.strip()]

                if len(headers) >= 2 and len(rows) >= 1:
                    table_type = self._determine_table_type(headers, rows)
                    logger.debug(f"Generic table parsing detected type: {table_type}")

                    if table_type in ['soil', 'leaf']:
                        if table_type == 'soil':
                            table_data = self._structure_soil_data(headers, rows)
                        else:
                            table_data = self._structure_leaf_data(headers, rows)

                        if table_data:
                            tables.append(table_data)

            return tables

        except Exception as e:
            logger.error(f"Generic table parsing failed: {e}")
            return []

    def _parse_delimited_text(self, text: str, delimiter: str) -> List[Optional[Dict]]:
        """Parse delimited text (CSV/TSV)"""
        try:
            tables = []
            lines = [line.strip() for line in text.split('\n') if line.strip()]

            for line in lines:
                if delimiter in line:
                    # This looks like a data row
                    parts = [part.strip() for part in line.split(delimiter) if part.strip()]
                    if len(parts) >= 3:  # At least sample ID + 2 parameters
                        # Try to reconstruct table from this line
                        headers = ['Sample ID'] + [f'Parameter_{i}' for i in range(len(parts)-1)]
                        rows = [parts]

                        table_type = self._determine_table_type(headers, rows)
                        if table_type in ['soil', 'leaf']:
                            if table_type == 'soil':
                                table_data = self._structure_soil_data(headers, rows)
                            else:
                                table_data = self._structure_leaf_data(headers, rows)

                            if table_data:
                                tables.append(table_data)

            return tables
            
        except Exception as e:
            logger.error(f"Delimited text parsing failed: {e}")
            return []
    
    def _parse_space_aligned_text(self, lines: List[str]) -> List[Optional[Dict]]:
        """Parse space-aligned tabular text"""
        try:
            tables = []

            # Look for lines that might be data rows
            data_lines = []
            for line in lines:
                # Check if line contains numbers and might be data
                import re
                if re.search(r'\d', line) and len(line.split()) >= 3:
                    data_lines.append(line)

            if len(data_lines) >= 2:  # At least 2 data lines
                # Try to parse as aligned columns
                # This is a simplified approach - in practice you'd need more sophisticated column detection
                headers = ['Sample ID'] + [f'Parameter_{i}' for i in range(len(data_lines[0].split())-1)]
                rows = [line.split() for line in data_lines]

                table_type = self._determine_table_type(headers, rows)
                if table_type in ['soil', 'leaf']:
                    if table_type == 'soil':
                        table_data = self._structure_soil_data(headers, rows)
                    else:
                        table_data = self._structure_leaf_data(headers, rows)

                    if table_data:
                        tables.append(table_data)

            return tables

        except Exception as e:
            logger.error(f"Space-aligned text parsing failed: {e}")
            return []

    def _parse_structured_sample_text(self, lines: List[str]) -> List[Optional[Dict]]:
        """Parse text that contains sample IDs and parameter values"""
        try:
            import re
            tables = []

            # More flexible sample ID patterns
            sample_patterns = [
                re.compile(r'S\d+/?\d*'),  # S1, S1/1, etc.
                re.compile(r'Sample\s*\d+'),  # Sample 1, Sample 2, etc.
                re.compile(r'Lab\s*No\.?\s*\d+'),  # Lab No 1, Lab No. 123, etc.
                re.compile(r'Farm\s*\d+'),  # Farm 1, Farm 2, etc.
                re.compile(r'Plot\s*\d+'),  # Plot 1, Plot 2, etc.
                re.compile(r'\b\d{1,3}\b'),  # Simple numbers 1-999
            ]

            sample_lines = []

            for line in lines:
                # Check if line contains any sample pattern
                for pattern in sample_patterns:
                    if pattern.search(line):
                        sample_lines.append(line)
                        break

            if len(sample_lines) >= 2:
                # Try to extract structured data
                headers = ['Sample ID', 'pH', 'N (%)', 'Org. C (%)', 'Total P (mg/kg)', 'Avail P (mg/kg)', 'Exch. K (meq%)', 'Exch. Ca (meq%)', 'Exch. Mg (meq%)', 'CEC (meq%)']
                rows = []

                for line in sample_lines:
                    # Extract sample ID and values using first matching pattern
                    sample_id = None
                    for pattern in sample_patterns:
                        match = pattern.search(line)
                        if match:
                            sample_id = match.group()
                            break

                    if sample_id:
                        # Extract numeric values from the line
                        values = re.findall(r'\d+\.?\d*', line)
                        if len(values) >= 3:  # At least sample ID + a few parameters
                            # Remove the sample ID number from values if it's there
                            sample_num_match = re.search(r'\d+', sample_id)
                            if sample_num_match:
                                sample_num = sample_num_match.group()
                                # Filter out the sample number from values if present
                                filtered_values = [v for v in values if v != sample_num]
                                if len(filtered_values) >= 3:
                                    rows.append([sample_id] + filtered_values[:9])
                                else:
                                    # If filtering removed too many, use original values
                                    rows.append([sample_id] + values[:9])

                if rows:
                    table_type = self._determine_table_type(headers, rows)
                    logger.info(f"Parsed structured text - detected type: {table_type}, rows: {len(rows)}")

                    if table_type in ['soil', 'leaf']:
                        if table_type == 'soil':
                            table_data = self._structure_soil_data(headers, rows)
                        else:
                            table_data = self._structure_leaf_data(headers, rows)

                        if table_data and table_data.get('samples'):
                            logger.info(f"Successfully structured {len(table_data['samples'])} samples from text parsing")
                            tables.append(table_data)
                        else:
                            logger.debug("No valid samples extracted from structured text parsing")
                    else:
                        logger.debug(f"Text parsing detected unsupported table type: {table_type}")

            return tables
            
        except Exception as e:
            logger.error(f"Structured sample text parsing failed: {e}")
            return []
    
    def _parse_text_for_tables(self, document_text: str) -> List[Optional[Dict]]:
        """Parse entire document text for tabular data"""
        try:
            tables = []

            # Split text into sections using multiple delimiters
            sections = re.split(r'\n\s*\n|\n\n|\n\s{2,}', document_text)

            logger.info(f"Split document into {len(sections)} sections for table parsing")

            for i, section in enumerate(sections):
                section = section.strip()
                if len(section) > 50:  # Process sections with meaningful content
                    logger.debug(f"Processing section {i+1}/{len(sections)} (length: {len(section)})")
                    table_data = self._parse_tabular_text_block(section)
                    if table_data:
                        logger.info(f"Found table in section {i+1}: {table_data.get('type', 'unknown')} with {len(table_data.get('samples', []))} samples")
                        tables.append(table_data)

            # If no tables found, try parsing the entire document as one block
            if not tables and len(document_text.strip()) > 200:
                logger.info("No tables found in sections, trying to parse entire document")
                table_data = self._parse_tabular_text_block(document_text)
                if table_data:
                    tables.append(table_data)

            logger.info(f"Total tables extracted from text parsing: {len(tables)}")
            return tables

        except Exception as e:
            logger.error(f"Document text parsing failed: {e}")
            return []

    def _build_markdown_table(self, title: str, headers: List[str], rows: List[List[Any]]) -> str:
        """Build a GitHub-flavored markdown table string from headers and rows."""
        if not headers or not rows:
            return ""

        header_line = "| " + " | ".join(headers) + " |"
        separator_line = "|" + "|".join(["-" * (len(h) + 2) for h in headers]) + "|"
        row_lines = []
        for row in rows:
            safe_row = [str(cell) if cell is not None else "" for cell in row]
            row_lines.append("| " + " | ".join(safe_row) + " |")

        table_md = f"### {title}\n" + header_line + "\n" + separator_line + "\n" + "\n".join(row_lines)
        return table_md

    def _tables_to_structured_sections(self, tables: List[Dict]) -> Dict[str, Any]:
        """Convert extracted tables to structured sections and markdown for preview."""
        sections: Dict[str, Any] = {
            'markdown': "",
            'tables': [],
            'dataframes': {}
        }

        soil_count = 0
        leaf_count = 0
        markdown_parts: List[str] = []

        for table in tables or []:
            table_type = table.get('type', 'unknown')
            headers = table.get('headers') or []
            samples = table.get('samples') or []

            # Reconstruct rows from structured samples
            rows: List[List[Any]] = []
            if table_type == 'soil':
                # Build default soil headers if missing
                if not headers:
                    headers = [
                        'Sample ID', 'pH', 'N (%)', 'Org. C (%)', 'Total P (mg/kg)', 'Avail P (mg/kg)',
                        'Exch. K (meq%)', 'Exch. Ca (meq%)', 'Exch. Mg (meq%)', 'CEC (meq%)'
                    ]

                # Normalize rows
                for sample in samples:
                    if isinstance(sample, dict) and 'sample_id' in sample and 'data' in sample:
                        data = sample['data']
                        row = [sample['sample_id']]
                        # Follow header order for remaining columns
                        for col in headers[1:]:
                            row.append(data.get(col))
                        rows.append(row)

                soil_count += 1
                md = self._build_markdown_table('Farm 3 Soil Test Data', headers or ['Sample ID'], rows)
                if md:
                    markdown_parts.append(md)
                # Also build a DataFrame for downstream use
                try:
                    df = pd.DataFrame(rows, columns=headers)
                    sections['dataframes']['soil'] = df
                except Exception:
                    pass
                sections['tables'].append({'title': 'Farm 3 Soil Test Data', 'headers': headers, 'rows': rows})

            elif table_type == 'leaf':
                # Flatten leaf structure to the requested format
                # Determine headers from observed data
                preferred_headers = [
                    'Sample ID', 'N (%)', 'P (%)', 'K (%)', 'Mg (%)', 'Ca (%)', 'B (mg/kg)', 'Cu (mg/kg)', 'Zn (mg/kg)'
                ]

                # Build dynamic headers from samples if needed
                headers = preferred_headers
                rows = []
                for sample in samples:
                    if isinstance(sample, dict):
                        sample_id = sample.get('sample_id') or sample.get('id') or sample.get('sample id')
                        n = sample.get('% Dry Matter', {}).get('N') or sample.get('N (%)')
                        p = sample.get('% Dry Matter', {}).get('P') or sample.get('P (%)')
                        k = sample.get('% Dry Matter', {}).get('K') or sample.get('K (%)')
                        mg = sample.get('% Dry Matter', {}).get('MG') or sample.get('% Dry Matter', {}).get('Mg') or sample.get('Mg (%)')
                        ca = sample.get('% Dry Matter', {}).get('CA') or sample.get('% Dry Matter', {}).get('Ca') or sample.get('Ca (%)')
                        b = sample.get('mg/kg Dry Matter', {}).get('B') or sample.get('B (mg/kg)')
                        cu = sample.get('mg/kg Dry Matter', {}).get('CU') or sample.get('Cu (mg/kg)')
                        zn = sample.get('mg/kg Dry Matter', {}).get('ZN') or sample.get('Zn (mg/kg)')
                        row = [sample_id, n, p, k, mg, ca, b, cu, zn]
                        rows.append(row)

                leaf_count += 1
                md = self._build_markdown_table('Farm 3 Leaf Test Data', headers, rows)
                if md:
                    markdown_parts.append(md)
                try:
                    df = pd.DataFrame(rows, columns=headers)
                    sections['dataframes']['leaf'] = df
                except Exception:
                    pass
                sections['tables'].append({'title': 'Farm 3 Leaf Test Data', 'headers': headers, 'rows': rows})

            else:
                # Unknown tables: attempt generic conversion
                gen_headers = ['Sample ID']
                # Infer columns from first sample
                first = next((s for s in samples if isinstance(s, dict)), None)
                if first:
                    keys = []
                    if 'data' in first and isinstance(first['data'], dict):
                        keys = list(first['data'].keys())
                    else:
                        keys = [k for k in first.keys() if k != 'sample_id']
                    gen_headers = ['Sample ID'] + keys

                gen_rows: List[List[Any]] = []
                for sample in samples:
                    if isinstance(sample, dict) and 'sample_id' in sample:
                        row = [sample['sample_id']]
                        if 'data' in sample and isinstance(sample['data'], dict):
                            for col in gen_headers[1:]:
                                row.append(sample['data'].get(col))
                        else:
                            for col in gen_headers[1:]:
                                row.append(sample.get(col))
                        gen_rows.append(row)

                md = self._build_markdown_table('Detected Table', gen_headers, gen_rows)
                if md:
                    markdown_parts.append(md)
                try:
                    df = pd.DataFrame(gen_rows, columns=gen_headers)
                    sections['dataframes']['generic'] = df
                except Exception:
                    pass
                sections['tables'].append({'title': 'Detected Table', 'headers': gen_headers, 'rows': gen_rows})

        sections['markdown'] = "\n\n".join([part for part in markdown_parts if part])
        return sections

    def _get_text_from_layout(self, layout, document_text: str) -> str:
        """Extract text from layout segments"""
        try:
            response = ""
            for segment in layout.text_anchor.text_segments:
                start_index = int(segment.start_index) if segment.start_index else 0
                end_index = int(segment.end_index) if segment.end_index else len(document_text)
                response += document_text[start_index:end_index]
            return response
        except Exception:
            return ""
    
    def _determine_table_type(self, headers: List[str], rows: List[List[str]]) -> str:
        """Determine if table contains soil or leaf analysis data"""
        header_text = ' '.join(headers).lower()
        row_text = ' '.join([' '.join(row) for row in rows[:3]]).lower()  # Check first few rows

        # Expanded soil indicators with more variations
        soil_indicators = [
            'ph', 'pH', 'ph value', 'soil ph',
            'organic carbon', 'org.c', 'org c', 'org. c', 'org c (%)', 'organic c',
            'cec', 'c.e.c', 'cec (meq%)', 'c.e.c (meq%)', 'cec meq%', 'cation exchange capacity',
            'exchangeable', 'exch.', 'exch k', 'exch ca', 'exch mg', 'exchangeable k', 'exchangeable ca', 'exchangeable mg',
            'avail p', 'available p', 'available phosphorus', 'p available', 'phosphorus available',
            'total p', 'total phosphorus', 'phosphorus total',
            'nitrogen', 'n (%)', 'n%', 'nitrogen (%)',
            'potassium', 'calcium', 'magnesium', 'phosphorus'
        ]

        # Expanded leaf indicators with more variations
        leaf_indicators = [
            '% dry matter', 'mg/kg dry matter', 'dry matter',
            'n (%)', 'p (%)', 'k (%)', 'mg (%)', 'ca (%)', 'n%', 'p%', 'k%', 'mg%', 'ca%',
            'nitrogen %', 'phosphorus %', 'potassium %', 'magnesium %', 'calcium %',
            'b (mg/kg)', 'cu (mg/kg)', 'zn (mg/kg)', 'boron', 'copper', 'zinc',
            'leaf analysis', 'foliar analysis', 'plant tissue', 'nutrient content'
        ]

        # Check for sample ID patterns that are more common in soil/leaf reports
        sample_patterns = ['s1', 's2', 's3', 'sample', 'lab no', 'farm', 'plot']
        has_sample_pattern = any(pattern in header_text or pattern in row_text for pattern in sample_patterns)

        soil_score = sum(1 for indicator in soil_indicators if indicator in header_text or indicator in row_text)
        leaf_score = sum(1 for indicator in leaf_indicators if indicator in header_text or indicator in row_text)

        # Add bonus points for having sample patterns
        if has_sample_pattern:
            soil_score += 1
            leaf_score += 1

        # Add bonus for having numeric data patterns
        numeric_pattern = re.compile(r'\d+\.?\d*')
        numeric_count = len(numeric_pattern.findall(row_text))
        if numeric_count > 10:  # Likely contains measurement data
            soil_score += 1
            leaf_score += 1

        logger.debug(f"Table detection scores - Soil: {soil_score}, Leaf: {leaf_score}, Sample patterns: {has_sample_pattern}, Numeric values: {numeric_count}")

        if soil_score > leaf_score:
            return 'soil'
        elif leaf_score > soil_score:
            return 'leaf'
        elif soil_score > 0 or leaf_score > 0:
            # If we have some indicators but equal scores, default to soil (more common)
            return 'soil'
        else:
            return 'unknown'
    
    def _structure_soil_data(self, headers: List[str], rows: List[List[str]]) -> Dict:
        """Structure soil analysis data"""
        try:
            samples = []
            
            # Map common soil parameters (more specific patterns first to avoid false matches)
            soil_param_mapping = {
                'sample id': ['sample id', 'sample no', 'sample_id', 'id', 'sample', 'lab no', 'lab_no', 'lab number', 'farm', 'plot'],
                'lab_no': ['lab no', 'lab_no', 'lab no.', 'lab number'],
                'ph': ['ph', 'pH', 'ph value', 'soil ph', 'ph level'],
                'cec': ['cec (meq%)', 'cec', 'c.e.c', 'c.e.c (meq%)', 'cec meq%', 'c.e.c meq%', 'cation exchange capacity', 'cec meq/100g'],
                # Put nitrogen BEFORE organic_carbon to avoid false matches
                'nitrogen': ['n (%)', 'nitrogen', 'n%', 'n', 'nitrogen (%)', 'n content (%)', 'total n', 'total nitrogen'],
                'organic_carbon': ['org.c (%)', 'organic carbon', 'org c', 'org. c', 'org c (%)', 'organic c', 'org. carbon (%)', 'organic matter', 'o.m (%)', 'o.m'],
                'exchangeable_k': ['exch. k (meq%)', 'exch k', 'exchangeable k', 'k (meq%)', 'exch. k', 'k meq', 'potassium (meq%)', 'k meq%', 'exch k meq%'],
                'exchangeable_ca': ['exch. ca (meq%)', 'exch ca', 'exchangeable ca', 'ca (meq%)', 'exch. ca', 'ca meq', 'calcium (meq%)', 'ca meq%', 'exch ca meq%'],
                'exchangeable_mg': ['exch. mg (meq%)', 'exch mg', 'exchangeable mg', 'mg (meq%)', 'exch. mg', 'mg meq', 'magnesium (meq%)', 'mg meq%', 'exch mg meq%'],
                'total_p': ['total p (mg/kg)', 'total p', 'total phosphorus', 'phosphorus total (mg/kg)', 'p total (mg/kg)', 'p total', 'phosphorus total'],
                'available_p': ['avail p (mg/kg)', 'available p', 'avail p', 'available phosphorus', 'p available (mg/kg)', 'avail. p (mg/kg)', 'p avail', 'phosphorus available']
            }
            
            # Create header mapping with flexible matching
            header_map = {}
            for i, header in enumerate(headers):
                header_lower = header.lower().strip()

                for param, variations in soil_param_mapping.items():
                    # More flexible matching: check if any variation is contained in the header
                    # or if the header is contained in any variation
                    for var in variations:
                        var_lower = var.lower()
                        if var_lower in header_lower or header_lower in var_lower:
                            header_map[i] = param
                            break

                    if i in header_map:
                        break
            
            # Process each row as a sample
            for row in rows:
                if len(row) >= len(headers):
                    sample = {}

                    # First column should be sample ID
                    sample_id = None
                    if row and row[0].strip():
                        sample_id = row[0].strip()

                    if sample_id:
                        # Map parameters to proper names - support multiple formats
                        param_name_mapping = {
                            'ph': 'pH',
                            'nitrogen': 'N (%)',  # Default to compact format, can be overridden
                            'organic_carbon': 'Org. C (%)',  # Default to compact format
                            'total_p': 'Total P (mg/kg)',
                            'available_p': 'Avail P (mg/kg)',
                            'exchangeable_k': 'Exch. K (meq%)',
                            'exchangeable_ca': 'Exch. Ca (meq%)',
                            'exchangeable_mg': 'Exch. Mg (meq%)',
                            'cec': 'CEC (meq%)'  # Default to compact format
                        }

                        # Detect format preference based on headers
                        format_preference = self._detect_format_preference(headers)

                        # Adjust mapping based on detected format
                        if format_preference == 'expanded':
                            param_name_mapping.update({
                                'nitrogen': 'Nitrogen (%)',
                                'organic_carbon': 'Organic Carbon (%)',
                                'available_p': 'Available P (mg/kg)',
                                'total_p': 'Total P (mg/kg)',
                                'cec': 'C.E.C (meq%)'
                            })
                        elif format_preference == 'compact':
                            # Keep default compact format
                            pass

                        for i, value in enumerate(row[1:], 1):  # Skip first column (sample ID)
                            if i in header_map and value.strip():
                                param_key = header_map[i]

                                if param_key in param_name_mapping:
                                    # Clean and convert values
                                    clean_value = self._clean_numeric_value(value)
                                    # Ensure numeric values are properly typed
                                    if isinstance(clean_value, str) and clean_value.replace('.', '').replace('-', '').isdigit():
                                        clean_value = float(clean_value) if '.' in clean_value else int(clean_value)
                                    final_param_name = param_name_mapping[param_key]
                                    sample[final_param_name] = clean_value

                        if len(sample) > 0:  # Only add if sample has data beyond just sample ID
                            samples.append({
                                'sample_id': sample_id,
                                'data': sample
                            })
            
            return {
                'type': 'soil',
                'headers': headers,
                'samples': samples,
                'total_samples': len(samples)
            }
            
        except Exception as e:
            logger.error(f"Soil data structuring failed: {e}")
            return {'type': 'soil', 'samples': [], 'error': str(e)}

    def _detect_format_preference(self, headers: List[str]) -> str:
        """Detect whether to use compact or expanded parameter naming"""
        header_text = ' '.join(headers).lower()

        # Look for indicators of expanded format
        expanded_indicators = [
            'nitrogen (%)', 'organic carbon (%)', 'available p', 'c.e.c'
        ]

        # Look for indicators of compact format
        compact_indicators = [
            'n (%)', 'org. c (%)', 'avail p', 'cec'
        ]

        expanded_score = sum(1 for indicator in expanded_indicators if indicator in header_text)
        compact_score = sum(1 for indicator in compact_indicators if indicator in header_text)

        # If expanded indicators are found, prefer expanded format
        if expanded_score > compact_score:
            return 'expanded'
        else:
            return 'compact'
    
    def _structure_leaf_data(self, headers: List[str], rows: List[List[str]]) -> Dict:
        """Structure leaf analysis data"""
        try:
            samples = []
            
            # Map leaf parameters with more variations
            leaf_param_mapping = {
                'sample_id': ['sample id', 'sample no', 'sample_id', 'lab no', 'lab no.', 'sample', 'id', 'farm', 'plot'],
                'n_percent': ['n (%)', 'n%', 'nitrogen %', 'nitrogen', 'n content (%)', 'total n (%)'],
                'p_percent': ['p (%)', 'p%', 'phosphorus %', 'phosphorus', 'p content (%)', 'total p (%)'],
                'k_percent': ['k (%)', 'k%', 'potassium %', 'potassium', 'k content (%)', 'total k (%)'],
                'mg_percent': ['mg (%)', 'mg%', 'magnesium %', 'magnesium', 'mg content (%)', 'total mg (%)'],
                'ca_percent': ['ca (%)', 'ca%', 'calcium %', 'calcium', 'ca content (%)', 'total ca (%)'],
                'b_mgkg': ['b (mg/kg)', 'b mg/kg', 'boron', 'b mg/kg dry matter', 'boron (mg/kg)'],
                'cu_mgkg': ['cu (mg/kg)', 'cu mg/kg', 'copper', 'cu mg/kg dry matter', 'copper (mg/kg)'],
                'zn_mgkg': ['zn (mg/kg)', 'zn mg/kg', 'zinc', 'zn mg/kg dry matter', 'zinc (mg/kg)'],
                'fe_mgkg': ['fe (mg/kg)', 'fe mg/kg', 'iron', 'fe mg/kg dry matter', 'iron (mg/kg)'],
                'mn_mgkg': ['mn (mg/kg)', 'mn mg/kg', 'manganese', 'mn mg/kg dry matter', 'manganese (mg/kg)']
            }
            
            # Create header mapping
            header_map = {}
            for i, header in enumerate(headers):
                header_lower = header.lower().strip()
                for param, variations in leaf_param_mapping.items():
                    if any(var in header_lower for var in variations):
                        header_map[i] = param
                        break
            
            # Process each row as a sample
            for row in rows:
                if len(row) >= len(headers):
                    sample = {
                        '% Dry Matter': {},
                        'mg/kg Dry Matter': {}
                    }
                    basic_info = {}
                    
                    for i, value in enumerate(row):
                        if i in header_map and value.strip():
                            param = header_map[i]
                            clean_value = self._clean_numeric_value(value)
                            
                            # Categorize parameters
                            if param in ['n_percent', 'p_percent', 'k_percent', 'mg_percent', 'ca_percent']:
                                nutrient_key = param.split('_')[0].upper()
                                sample['% Dry Matter'][nutrient_key] = clean_value
                            elif param in ['b_mgkg', 'cu_mgkg', 'zn_mgkg']:
                                nutrient_key = param.split('_')[0].upper()
                                sample['mg/kg Dry Matter'][nutrient_key] = clean_value
                            else:
                                basic_info[param] = clean_value
                    
                    # Add basic info
                    sample.update(basic_info)
                    
                    if basic_info or sample['% Dry Matter'] or sample['mg/kg Dry Matter']:
                        samples.append(sample)
            
            return {
                'type': 'leaf',
                'headers': headers,
                'samples': samples,
                'total_samples': len(samples)
            }
            
        except Exception as e:
            logger.error(f"Leaf data structuring failed: {e}")
            return {'type': 'leaf', 'samples': [], 'error': str(e)}
    
    def _clean_numeric_value(self, value: str) -> Any:
        """Clean and convert numeric values"""
        if not value or not isinstance(value, str):
            return value

        # Remove common non-numeric characters but preserve decimal points
        cleaned = re.sub(r'[^\d.\-<>]', '', value.strip())

        # Handle special cases
        if 'n.d.' in value.lower() or 'nd' in value.lower():
            return 'N.D.'
        if '<' in value:
            return f"<{cleaned.replace('<', '')}"
        if '>' in value:
            return f">{cleaned.replace('>', '')}"
        
        # Try to convert to number
        try:
            if '.' in cleaned:
                return float(cleaned)
            else:
                return int(cleaned)
        except ValueError:
            return value  # Return original if conversion fails


class TesseractProcessor:
    """Fallback OCR processor using Tesseract"""
    
    def __init__(self):
        self.available = TESSERACT_AVAILABLE
        self._configure_tesseract_path()
    
    def _configure_tesseract_path(self):
        """Configure Tesseract executable path from secrets or environment"""
        if not TESSERACT_AVAILABLE:
            return
        
        try:
            import streamlit as st
            # Try to get path from secrets
            if hasattr(st, 'secrets') and st.secrets is not None:
                try:
                    tesseract_path = st.secrets.get('ocr', {}).get('tesseract_path')
                    if tesseract_path and os.path.exists(tesseract_path):
                        pytesseract.pytesseract.tesseract_cmd = tesseract_path
                        logger.info(f"Tesseract path configured from secrets: {tesseract_path}")
                        return
                except Exception as e:
                    logger.debug(f"Could not read tesseract_path from secrets: {e}")
        except ImportError:
            pass
        
        # Try common Windows installation paths
        common_paths = [
            r"C:\Program Files\Tesseract-OCR\tesseract.exe",
            r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
            r"C:\Users\{}\AppData\Local\Programs\Tesseract-OCR\tesseract.exe".format(os.getenv('USERNAME', '')),
        ]
        
        for path in common_paths:
            if os.path.exists(path):
                pytesseract.pytesseract.tesseract_cmd = path
                logger.info(f"Tesseract path auto-detected: {path}")
                return
        
        # Try to find tesseract in PATH
        try:
            import shutil
            tesseract_cmd = shutil.which('tesseract')
            if tesseract_cmd:
                pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
                logger.info(f"Tesseract found in PATH: {tesseract_cmd}")
        except Exception:
            pass
    
    def process_document(self, file_path: str) -> Optional[Dict]:
        """Process document with Tesseract OCR"""
        if not self.available:
            return None
        
        try:
            # Handle different file types
            if file_path.lower().endswith('.pdf'):
                images = self._pdf_to_images(file_path)
                if not images:
                    return None
                # Process first page for now
                image = images[0]
            else:
                image = Image.open(file_path)
            
            # Preprocess image for better OCR
            processed_image = self._preprocess_image(image)
            
            # Extract text using Tesseract
            text = pytesseract.image_to_string(processed_image, config='--psm 6')
            
            # Try to extract tabular data
            table_data = self._extract_table_from_text(text)
            
            return {
                'text': text,
                'tables': [table_data] if table_data else [],
                'success': True,
                'method': 'tesseract_fallback'
            }
            
        except Exception as e:
            logger.error(f"Tesseract processing failed: {e}")
            return None
    
    def _pdf_to_images(self, pdf_path: str) -> List[Image.Image]:
        """Convert PDF to images"""
        if not PDF_AVAILABLE:
            return []
        
        try:
            doc = fitz.open(pdf_path)
            images = []
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                mat = fitz.Matrix(2.0, 2.0)  # Increase resolution
                pix = page.get_pixmap(matrix=mat)
                img_data = pix.tobytes("ppm")
                image = Image.open(io.BytesIO(img_data))
                images.append(image)
            doc.close()
            return images
        except Exception as e:
            logger.error(f"PDF to image conversion failed: {e}")
            return []
    
    def _preprocess_image(self, image: Image.Image) -> Image.Image:
        """Preprocess image for better OCR results"""
        try:
            # Convert PIL to OpenCV format
            opencv_image = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
            
            # Convert to grayscale
            gray = cv2.cvtColor(opencv_image, cv2.COLOR_BGR2GRAY)
            
            # Apply threshold to get binary image
            _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
            # Remove noise
            kernel = np.ones((1, 1), np.uint8)
            binary = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel)
            binary = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel)
            
            # Convert back to PIL
            return Image.fromarray(binary)

        except Exception as e:
            logger.error(f"Image preprocessing failed: {e}")
            return image
    
    def _extract_table_from_text(self, text: str) -> Optional[Dict]:
        """Extract tabular data from OCR text using pattern matching"""
        try:
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            
            # Look for table patterns
            potential_rows = []
            headers = []
            
            for line in lines:
                # Split by multiple spaces or tabs to identify columns
                columns = re.split(r'\s{2,}|\t', line)
                if len(columns) > 3:  # Likely a table row
                    potential_rows.append(columns)
            
            if not potential_rows:
                return None

            # First row is likely headers
            headers = potential_rows[0]
            rows = potential_rows[1:] if len(potential_rows) > 1 else []
            
            # Determine table type
            table_type = self._determine_table_type_from_text(text)
            
            return {
                'type': table_type,
                'headers': headers,
                'rows': rows,
                'samples': self._structure_data_from_text(table_type, headers, rows)
            }
            
        except Exception as e:
            logger.error(f"Text table extraction failed: {e}")
            return None

    def _determine_table_type_from_text(self, text: str) -> str:
        """Determine table type from text content"""
        text_lower = text.lower()
        
        soil_indicators = ['ph', 'organic carbon', 'cec', 'exchangeable', 'available p']
        leaf_indicators = ['% dry matter', 'mg/kg dry', 'n (%)', 'p (%)', 'k (%)']
        
        soil_score = sum(1 for indicator in soil_indicators if indicator in text_lower)
        leaf_score = sum(1 for indicator in leaf_indicators if indicator in text_lower)
        
        return 'soil' if soil_score > leaf_score else 'leaf' if leaf_score > 0 else 'unknown'
    
    def _structure_data_from_text(self, table_type: str, headers: List[str], rows: List[List[str]]) -> List[Dict]:
        """Structure data based on table type"""
        samples = []
        
        for row in rows:
            if len(row) > 1:
                sample = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        key = headers[i].strip()
                        sample[key] = value.strip()
                samples.append(sample)
        
        return samples


def _detect_report_type(sample_ids: List[str]) -> str:
    """Detect the report type based on sample ID patterns"""
    if not sample_ids:
        return "Farm_3_Soil_Test_Data"  # Default

    # Check patterns
    sample_patterns = {
        "Farm_3_Soil_Test_Data": [sid for sid in sample_ids if sid.startswith('S') and len(sid) == 4 and sid[1:].isdigit()],
        "SP_Lab_Test_Report": [sid for sid in sample_ids if '/' in sid and 'S' in sid]
    }

    # Count matches for each pattern
    farm_count = len(sample_patterns["Farm_3_Soil_Test_Data"])
    sp_count = len(sample_patterns["SP_Lab_Test_Report"])

    # Return the pattern with more matches
    if sp_count > farm_count:
        return "SP_Lab_Test_Report"
    else:
        return "Farm_3_Soil_Test_Data"


def _process_csv_file(csv_path: str) -> Optional[Dict]:
    """Process CSV file directly for table extraction"""
    try:
        import csv

        logger.info(f"Processing CSV file: {os.path.basename(csv_path)}")

        # Read CSV/text file
        with open(csv_path, 'r', encoding='utf-8') as file:
            content = file.read()

        # Try to detect delimiter
        delimiter = ','
        if '\t' in content:
            delimiter = '\t'
        elif '|' in content:
            delimiter = '|'
        elif ';' in content:
            delimiter = ';'

        # Split into rows and then into columns
        rows = []
        for line in content.split('\n'):
            line = line.strip()
            if line:  # Skip empty lines
                # Split by detected delimiter
                parts = [part.strip() for part in line.split(delimiter) if part.strip()]
                if parts:  # Only add non-empty rows
                    rows.append(parts)

        if not rows or len(rows) < 2:
            logger.warning("CSV file has insufficient data")
            return None

        # First row is headers
        headers = [cell.strip() for cell in rows[0]]
        data_rows = rows[1:]

        logger.info(f"CSV parsed: {len(headers)} headers, {len(data_rows)} data rows")

        # Process the data
        table_data = _extract_table_data_from_csv(headers, data_rows)

        result = {
            'success': True,
            'tables': [table_data] if table_data else [],
            'text': f"CSV file with {len(data_rows)} rows",
            'raw_data': {
                'headers': headers,
                'rows': data_rows,
                'extraction_details': {}
            }
        }

        # Format the output in the requested structure
        if table_data and table_data.get('type') == 'soil':
            logger.debug(f"Processing soil table with {table_data.get('total_samples', 0)} samples")

            soil_samples = {}
            for sample in table_data.get('samples', []):
                if isinstance(sample, dict) and 'sample_id' in sample and 'data' in sample:
                    sample_id = sample['sample_id']
                    sample_data = sample['data']
                    soil_samples[sample_id] = sample_data
                    logger.debug(f"Added sample {sample_id} with parameters: {list(sample_data.keys())}")

            if soil_samples:
                # Detect the report type based on sample naming pattern
                report_type = _detect_report_type(list(soil_samples.keys()))

                result['raw_data']['extraction_details'][report_type] = soil_samples
                logger.debug(f"Created {report_type} with {len(soil_samples)} samples")

                result['raw_data']['extraction_details']['soil_summary'] = {
                    'total_samples': len(soil_samples),
                    'sample_ids': list(soil_samples.keys()),
                    'parameters': list(set([param for sample_data in soil_samples.values() for param in sample_data.keys()])),
                    'report_type': report_type
                }

        return result

    except Exception as e:
        logger.error(f"CSV processing failed: {e}")
        return None


def _process_excel_file(excel_path: str) -> Optional[Dict]:
    """Process Excel file directly for table extraction"""
    if not EXCEL_AVAILABLE:
        logger.error("Excel processing libraries not available")
        return None
        
    try:
        import pandas as pd
        from openpyxl import load_workbook
        import xlrd

        logger.info(f"Processing Excel file: {os.path.basename(excel_path)}")

        # Determine file type and read accordingly
        file_ext = os.path.splitext(excel_path)[1].lower()

        if file_ext == '.xlsx':
            # Use openpyxl for .xlsx files
            workbook = load_workbook(excel_path, data_only=True)
            sheet = workbook.active
            data = []

            # Read all rows
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):  # Skip completely empty rows
                    data.append([str(cell) if cell is not None else '' for cell in row])

        elif file_ext == '.xls':
            # Use xlrd for .xls files
            workbook = xlrd.open_workbook(excel_path)
            sheet = workbook.sheet_by_index(0)
            data = []

            # Read all rows
            for row_idx in range(sheet.nrows):
                row = []
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    row.append(str(cell_value) if cell_value is not None else '')
                if any(cell.strip() for cell in row):  # Skip rows with only empty strings
                    data.append(row)

        if not data or len(data) < 2:
            logger.warning("Excel file has insufficient data")
            return None

        # First row is headers
        headers = [cell.strip() for cell in data[0]]
        data_rows = data[1:]

        # Filter out empty rows from data
        data_rows = [row for row in data_rows if any(cell.strip() for cell in row)]

        logger.info(f"Excel parsed: {len(headers)} headers, {len(data_rows)} data rows")

        # Process the data
        logger.info(f"Processing Excel data with headers: {headers[:5]}...")  # Show first 5 headers
        logger.info(f"Sample data row: {data_rows[0] if data_rows else 'No data rows'}")
        
        table_data = _extract_table_data_from_excel(headers, data_rows)
        
        if table_data:
            logger.info(f"Successfully extracted table data: type={table_data.get('type')}, samples={len(table_data.get('samples', []))}")
        else:
            logger.warning("No table data extracted from Excel file")

        result = {
            'success': True,
            'tables': [table_data] if table_data else [],
            'text': f"Excel file with {len(data_rows)} rows",
            'raw_data': {
                'headers': headers,
                'rows': data_rows,
                'extraction_details': {}
            }
        }

        # Format the output in the requested structure
        if table_data and table_data.get('type') == 'soil':
            logger.debug(f"Processing soil table with {table_data.get('total_samples', 0)} samples")

            soil_samples = {}
            for sample in table_data.get('samples', []):
                if isinstance(sample, dict) and 'sample_id' in sample and 'data' in sample:
                    sample_id = sample['sample_id']
                    sample_data = sample['data']
                    soil_samples[sample_id] = sample_data
                    logger.debug(f"Added sample {sample_id} with parameters: {list(sample_data.keys())}")

            if soil_samples:
                # Detect the report type based on sample naming pattern
                report_type = _detect_report_type(list(soil_samples.keys()))

                result['raw_data']['extraction_details']['soil_samples'] = soil_samples
                result['raw_data']['extraction_details']['soil_summary'] = {
                    'total_samples': len(soil_samples),
                    'sample_ids': list(soil_samples.keys()),
                    'parameters': list(set([param for sample_data in soil_samples.values() for param in sample_data.keys()])),
                    'report_type': report_type
                }

        elif table_data and table_data.get('type') == 'leaf':
            logger.debug(f"Processing leaf table with {table_data.get('total_samples', 0)} samples")

            leaf_samples = {}
            for sample in table_data.get('samples', []):
                if isinstance(sample, dict) and 'sample_id' in sample and 'data' in sample:
                    sample_id = sample['sample_id']
                    sample_data = sample['data']
                    leaf_samples[sample_id] = sample_data
                    logger.debug(f"Added sample {sample_id} with parameters: {list(sample_data.keys())}")

            if leaf_samples:
                result['raw_data']['extraction_details']['leaf_samples'] = leaf_samples
                result['raw_data']['extraction_details']['leaf_summary'] = {
                    'total_samples': len(leaf_samples),
                    'sample_ids': list(leaf_samples.keys()),
                    'parameters': list(set([param for sample_data in leaf_samples.values() for param in sample_data.keys()]))
                }

        return result

    except ImportError as e:
        logger.error(f"Excel processing libraries not available: {e}")
        logger.error("Install required libraries: pip install openpyxl xlrd pandas")
        return None
    except Exception as e:
        logger.error(f"Excel processing failed: {e}")
        return None


def _extract_table_data_from_excel(headers: List[str], rows: List[List[str]]) -> Optional[Dict]:
    """Extract structured data from Excel table"""
    try:
        logger.info(f"Processing Excel table with {len(headers)} headers and {len(rows)} rows")
        logger.debug(f"Headers: {headers}")

        # Determine table type using class method
        processor = DocumentAIProcessor()
        table_type = processor._determine_table_type(headers, rows)
        logger.info(f"Detected table type: {table_type}")

        if table_type == 'soil':
            result = processor._structure_soil_data(headers, rows)
            logger.info(f"Soil data structured: {len(result.get('samples', []))} samples")
            return result
        elif table_type == 'leaf':
            result = processor._structure_leaf_data(headers, rows)
            logger.info(f"Leaf data structured: {len(result.get('samples', []))} samples")
            return result
        else:
            # Create a generic structure
            samples = []
            for row_idx, row in enumerate(rows):
                if len(row) >= len(headers):
                    sample = {}
                    for i, value in enumerate(row):
                        if i < len(headers) and headers[i].strip():
                            sample[headers[i].strip()] = str(value).strip() if value else ''
                    if sample and any(v for v in sample.values() if v):  # Only add non-empty samples
                        samples.append(sample)
                        
            logger.info(f"Generic structure created with {len(samples)} samples")
            return {
                'type': 'generic',
                'headers': headers,
                'samples': samples,
                'total_samples': len(samples)
            }

    except Exception as e:
        logger.error(f"Excel table extraction failed: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return None


def extract_data_from_image(image_path: str) -> Dict[str, Any]:
    """
    Main function to extract data from images using Google Document AI with Tesseract fallback
    
    Args:
        image_path (str): Path to the image file
        
    Returns:
        Dict containing extraction results with structured data
    """
    
    result = {
                'success': False,
        'error': None,
        'method': None,
        'tables': [],
        'raw_data': None,
        'extraction_details': {}
    }
    
    try:
        # Validate file exists
        if not os.path.exists(image_path):
            result['error'] = f"File not found: {image_path}"
            return result
        
        # Check if it's a CSV, Excel, or text file and handle differently
        file_ext = os.path.splitext(image_path)[1].lower()
        if file_ext in ['.csv', '.txt', '.tsv']:
            csv_result = _process_csv_file(image_path)
            if csv_result and csv_result.get('success'):
                result['success'] = True
                result['method'] = f'{file_ext[1:]}_parser'
                result['tables'] = csv_result.get('tables', [])
                result['raw_data'] = csv_result.get('raw_data', {})
                # Add structured markdown sections for preview
                try:
                    processor = DocumentAIProcessor()
                    sections = processor._tables_to_structured_sections(result['tables'])
                    if sections and sections.get('markdown'):
                        if 'extraction_details' not in result['raw_data']:
                            result['raw_data']['extraction_details'] = {}
                        result['raw_data']['structured_markdown'] = sections['markdown']
                        result['raw_data']['structured_tables'] = sections['tables']
                except Exception:
                    pass
                return result
        elif file_ext in ['.xlsx', '.xls']:
            if not EXCEL_AVAILABLE:
                result['error'] = "Excel processing libraries not available. Install with: pip install openpyxl xlrd"
                return result
                
            excel_result = _process_excel_file(image_path)
            if excel_result and excel_result.get('success'):
                result['success'] = True
                result['method'] = f'{file_ext[1:]}_parser'
                result['tables'] = excel_result.get('tables', [])
                result['raw_data'] = excel_result.get('raw_data', {})
                # Add structured markdown sections for preview
                try:
                    processor = DocumentAIProcessor()
                    sections = processor._tables_to_structured_sections(result['tables'])
                    if sections and sections.get('markdown'):
                        if 'extraction_details' not in result['raw_data']:
                            result['raw_data']['extraction_details'] = {}
                        result['raw_data']['structured_markdown'] = sections['markdown']
                        result['raw_data']['structured_tables'] = sections['tables']
                except Exception:
                    pass
                return result

        # Try Google Document AI first (but skip for Excel files as they're not supported)
        if DOCUMENT_AI_AVAILABLE and file_ext not in ['.xlsx', '.xls']:
            doc_ai = DocumentAIProcessor()
            doc_result = doc_ai.process_document(image_path)
            
            if doc_result and doc_result.get('success'):
                result['success'] = True
                result['method'] = 'document_ai'
                result['tables'] = doc_result.get('tables', [])
                raw_text = doc_result.get('text', '')
                logger.info(f"Document AI extracted raw text: {len(raw_text)} characters")
                if raw_text:
                    logger.debug(f"Raw text preview: {raw_text[:200]}...")
                    logger.info(f"Raw text contains sample IDs: {any('S00' in raw_text or 'L00' in raw_text for pattern in ['S00', 'L00'])}")
                    logger.info(f"Raw text contains soil keywords: {any(keyword in raw_text.lower() for keyword in ['cec', 'ph', 'organic'])}")

                result['raw_data'] = {
                    'text': raw_text,
                    'extraction_details': {}
                }
                
                # Structure the extracted data for each table
                for table in result['tables']:
                    table_type = table.get('type', 'unknown')
                    if table_type == 'soil':
                        soil_samples = {}
                        for sample in table.get('samples', []):
                            if isinstance(sample, dict) and 'sample_id' in sample and 'data' in sample:
                                sample_id = sample['sample_id']
                                soil_samples[sample_id] = sample['data']

                        if soil_samples:
                            # Detect the report type based on sample naming pattern
                            report_type = _detect_report_type(list(soil_samples.keys()))

                            result['raw_data']['extraction_details'][report_type] = soil_samples

                        result['raw_data']['extraction_details']['soil_summary'] = {
                            'total_samples': len(soil_samples),
                            'sample_ids': list(soil_samples.keys()),
                            'parameters': list(set([param for sample_data in soil_samples.values() for param in sample_data.keys()])),
                            'report_type': report_type if 'report_type' in locals() else 'Farm_3_Soil_Test_Data'
                        }
                    elif table_type == 'leaf':
                        result['raw_data']['extraction_details']['leaf_data'] = {
                            'samples': table.get('samples', []),
                            'total_samples': table.get('total_samples', 0),
                            'parameters': list(set([key for sample in table.get('samples', []) for key in sample.keys() if key not in ['% Dry Matter', 'mg/kg Dry Matter']]))
                        }
                
                # Add structured markdown sections for preview
                try:
                    sections = doc_ai._tables_to_structured_sections(result['tables']) if hasattr(doc_ai, '_tables_to_structured_sections') else self._tables_to_structured_sections(result['tables'])
                    if sections and sections.get('markdown'):
                        result['raw_data']['structured_markdown'] = sections['markdown']
                        result['raw_data']['structured_tables'] = sections['tables']
                except Exception as _:
                    pass

                # If no tables found, try to parse the raw text for structured data
                if not result['tables'] and result['raw_data'] and result['raw_data'].get('text'):
                    raw_text = result['raw_data']['text']
                    logger.info("No tables found, attempting to parse raw text for structured data")
                    logger.info(f"Raw text to parse: {len(raw_text)} characters")
                    logger.debug(f"Raw text content: {raw_text[:500]}...")

                    try:
                        # Use the existing DocumentAI processor to parse text for tables
                        parsed_tables = doc_ai._parse_text_for_tables(raw_text)

                        if parsed_tables:
                            logger.info(f"Found {len(parsed_tables)} tables from raw text parsing")
                            result['tables'] = parsed_tables
                        else:
                            logger.debug("Raw text parsing returned no tables - this is normal for some document types")

                            # Update extraction details with parsed data
                            for table in parsed_tables:
                                table_type = table.get('type', 'unknown')
                                if table_type == 'soil':
                                    soil_samples = {}
                                    for sample in table.get('samples', []):
                                        if isinstance(sample, dict) and 'Sample ID' in sample:
                                            sample_id = sample['Sample ID']
                                            sample_data = {k: v for k, v in sample.items() if k != 'Sample ID'}
                                            soil_samples[sample_id] = sample_data

                                    if soil_samples:
                                        report_type = _detect_report_type(list(soil_samples.keys()))
                                        if 'extraction_details' not in result['raw_data']:
                                            result['raw_data']['extraction_details'] = {}
                                        result['raw_data']['extraction_details'][report_type] = soil_samples
                                        result['raw_data']['extraction_details']['soil_summary'] = {
                                            'total_samples': len(soil_samples),
                                            'sample_ids': list(soil_samples.keys()),
                                            'parameters': list(set([param for sample_data in soil_samples.values() for param in sample_data.keys()])),
                                            'report_type': report_type
                                        }
                                elif table_type == 'leaf':
                                    if 'extraction_details' not in result['raw_data']:
                                        result['raw_data']['extraction_details'] = {}
                                    result['raw_data']['extraction_details']['leaf_data'] = {
                                        'samples': table.get('samples', []),
                                        'total_samples': len(table.get('samples', [])),
                                        'parameters': list(set([key for sample in table.get('samples', []) for key in sample.keys() if key != 'Sample ID']))
                                    }
                    except Exception as e:
                        logger.warning(f"Failed to parse raw text for structured data: {e}")

                logger.info(f"Document AI extraction successful: {len(result['tables'])} tables found")
                if result['tables']:
                    for i, table in enumerate(result['tables']):
                        logger.info(f"Table {i+1}: Type={table.get('type', 'unknown')}, Samples={len(table.get('samples', []))}")
                return result
        
        # Fallback to Tesseract if Document AI fails or is not available
        if TESSERACT_AVAILABLE:
            tesseract = TesseractProcessor()
            tess_result = tesseract.process_document(image_path)
            
            if tess_result and tess_result.get('success'):
                result['success'] = True
                result['method'] = 'tesseract_fallback'
                result['tables'] = tess_result.get('tables', [])
                raw_text = tess_result.get('text', '')
                logger.info(f"Tesseract extracted raw text: {len(raw_text)} characters")
                if raw_text:
                    logger.debug(f"Raw text preview: {raw_text[:200]}...")
                    logger.info(f"Raw text contains sample IDs: {any('S00' in raw_text or 'L00' in raw_text for pattern in ['S00', 'L00'])}")
                    logger.info(f"Raw text contains soil keywords: {any(keyword in raw_text.lower() for keyword in ['cec', 'ph', 'organic'])}")

                result['raw_data'] = {
                    'text': raw_text,
                    'extraction_details': {}
                }

                # If no tables found, try to parse the raw text for structured data
                if not result['tables'] and result['raw_data'] and result['raw_data'].get('text'):
                    logger.info("No tables found, attempting to parse raw text for structured data")
                    try:
                        # Use the existing DocumentAI processor to parse text for tables
                        parsed_tables = doc_ai._parse_text_for_tables(raw_text)

                        if parsed_tables:
                            logger.info(f"Found {len(parsed_tables)} tables from raw text parsing")
                            result['tables'] = parsed_tables

                            # Update extraction details with parsed data
                            for table in parsed_tables:
                                table_type = table.get('type', 'unknown')
                                if table_type == 'soil':
                                    soil_samples = {}
                                    for sample in table.get('samples', []):
                                        if isinstance(sample, dict) and 'Sample ID' in sample:
                                            sample_id = sample['Sample ID']
                                            sample_data = {k: v for k, v in sample.items() if k != 'Sample ID'}
                                            soil_samples[sample_id] = sample_data

                                    if soil_samples:
                                        report_type = _detect_report_type(list(soil_samples.keys()))
                                        if 'extraction_details' not in result['raw_data']:
                                            result['raw_data']['extraction_details'] = {}
                                        result['raw_data']['extraction_details'][report_type] = soil_samples
                                        result['raw_data']['extraction_details']['soil_summary'] = {
                                            'total_samples': len(soil_samples),
                                            'sample_ids': list(soil_samples.keys()),
                                            'parameters': list(set([param for sample_data in soil_samples.values() for param in sample_data.keys()])),
                                            'report_type': report_type
                                        }
                                elif table_type == 'leaf':
                                    if 'extraction_details' not in result['raw_data']:
                                        result['raw_data']['extraction_details'] = {}
                                    result['raw_data']['extraction_details']['leaf_data'] = {
                                        'samples': table.get('samples', []),
                                        'total_samples': len(table.get('samples', [])),
                                        'parameters': list(set([key for sample in table.get('samples', []) for key in sample.keys() if key != 'Sample ID']))
                                    }
                    except Exception as e:
                        logger.warning(f"Failed to parse raw text for structured data: {e}")

                # Structure the extracted data
                for table in result['tables']:
                    table_type = table.get('type', 'unknown')
                    if table_type == 'soil':
                        soil_samples = {}
                        for sample in table.get('samples', []):
                            if isinstance(sample, dict) and 'sample_id' in sample and 'data' in sample:
                                sample_id = sample['sample_id']
                                soil_samples[sample_id] = sample['data']

                        if soil_samples:
                            # Detect the report type based on sample naming pattern
                            report_type = _detect_report_type(list(soil_samples.keys()))

                            result['raw_data']['extraction_details'][report_type] = soil_samples

                        result['raw_data']['extraction_details']['soil_summary'] = {
                            'total_samples': len(soil_samples),
                            'sample_ids': list(soil_samples.keys()),
                            'parameters': list(set([param for sample_data in soil_samples.values() for param in sample_data.keys()])),
                            'report_type': report_type if 'report_type' in locals() else 'Farm_3_Soil_Test_Data'
                        }
                    elif table_type == 'leaf':
                        result['raw_data']['extraction_details']['leaf_data'] = {
                            'samples': table.get('samples', []),
                            'total_samples': len(table.get('samples', [])),
                            'parameters': list(set([key for sample in table.get('samples', []) for key in sample.keys()]))
                        }
                
                # Add structured markdown sections for preview
                try:
                    sections = self._tables_to_structured_sections(result['tables'])
                    if sections and sections.get('markdown'):
                        result['raw_data']['structured_markdown'] = sections['markdown']
                        result['raw_data']['structured_tables'] = sections['tables']
                except Exception as _:
                    pass

                logger.info(f"Tesseract extraction successful: {len(result['tables'])} tables found")
                return result
        
        # If both methods fail
        result['error'] = "No OCR methods available or all methods failed"
        if not DOCUMENT_AI_AVAILABLE and not TESSERACT_AVAILABLE:
            result['error'] = "No OCR libraries installed. Please install google-cloud-documentai or pytesseract"
        
        return result
        
    except Exception as e:
        logger.error(f"OCR extraction failed: {e}")
        result['error'] = f"Extraction error: {str(e)}"
        return result


# Utility functions for data validation and cleaning
def validate_soil_data(soil_samples: List[Dict]) -> Dict[str, Any]:
    """Validate extracted soil data"""
    validation_result = {
        'is_valid': True,
        'issues': [],
        'recommendations': []
    }
    
    # Map parameter names to their possible variations
    required_param_mapping = {
        'ph': ['pH', 'ph', 'Ph', 'PH'],
        'organic_carbon': ['Org. C (%)', 'Organic Carbon (%)', 'organic_carbon', 'org_c', 'Org.C (%)', 'Org C (%)'],
        'available_p': ['Avail P (mg/kg)', 'Available P (mg/kg)', 'available_p', 'avail_p', 'Avail. P (mg/kg)']
    }
    
    for i, sample in enumerate(soil_samples):
        sample_issues = []
        
        # Get sample data - handle both flat and nested structures
        sample_data = sample.get('data', sample) if 'data' in sample else sample
        
        # Check for required parameters
        missing_params = []
        for param_key, param_variations in required_param_mapping.items():
            found = False
            for variation in param_variations:
                if variation in sample_data:
                    found = True
                    break
            if not found:
                missing_params.append(param_key)
        
        if missing_params:
            sample_issues.append(f"Missing parameters: {', '.join(missing_params)}")
        
        # Validate pH range
        ph_value = None
        for ph_variation in required_param_mapping['ph']:
            if ph_variation in sample_data:
                try:
                    ph_value = float(sample_data[ph_variation])
                    break
                except (ValueError, TypeError):
                    continue
        
        if ph_value is not None:
            if ph_value < 3 or ph_value > 10:
                sample_issues.append(f"pH value {ph_value} seems out of normal range (3-10)")
        
        if sample_issues:
            validation_result['issues'].append(f"Sample {i+1}: {'; '.join(sample_issues)}")
            validation_result['is_valid'] = False
    
    return validation_result


def validate_leaf_data(leaf_samples: List[Dict]) -> Dict[str, Any]:
    """Validate extracted leaf data"""
    validation_result = {
        'is_valid': True,
        'issues': [],
        'recommendations': []
    }
    
    # Map nutrient names to their possible variations
    required_nutrient_mapping = {
        'N': ['N (%)', 'Nitrogen (%)', 'N', 'nitrogen', 'N%'],
        'P': ['P (%)', 'Phosphorus (%)', 'P', 'phosphorus', 'P%'],
        'K': ['K (%)', 'Potassium (%)', 'K', 'potassium', 'K%']
    }
    
    for i, sample in enumerate(leaf_samples):
        sample_issues = []
        
        # Get sample data - handle both flat and nested structures
        sample_data = sample.get('data', sample) if 'data' in sample else sample
        
        # Check for required nutrients
        missing_nutrients = []
        for nutrient_key, nutrient_variations in required_nutrient_mapping.items():
            found = False
            
            # Check in main data
            for variation in nutrient_variations:
                if variation in sample_data:
                    found = True
                    break
            
            # Also check in % Dry Matter section if it exists
            if not found and '% Dry Matter' in sample_data:
                dry_matter = sample_data['% Dry Matter']
                for variation in nutrient_variations:
                    if variation in dry_matter:
                        found = True
                        break
            
            if not found:
                missing_nutrients.append(nutrient_key)
        
        if missing_nutrients:
            sample_issues.append(f"Missing nutrients: {', '.join(missing_nutrients)}")
        
        if sample_issues:
            validation_result['issues'].append(f"Sample {i+1}: {'; '.join(sample_issues)}")
            validation_result['is_valid'] = False
    
    return validation_result


if __name__ == "__main__":
    # Test the extraction function
    import sys
    if len(sys.argv) > 1:
        test_file = sys.argv[1]
        result = extract_data_from_image(test_file)
        print(json.dumps(result, indent=2, default=str))